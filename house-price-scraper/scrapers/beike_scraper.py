# -*- coding: utf-8 -*-
"""
贝壳网爬虫 - 成都二手房数据爬取

用法:
    python scrapers/beike_scraper.py <小区名> [几室]
示例:
    python scrapers/beike_scraper.py 鑫苑鑫都汇 3
    python scrapers/beike_scraper.py 鑫苑鑫都汇 三室二厅
"""

import sys, os, re, time, random, urllib.request
from io import BytesIO
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import config
from scrapers.driver_utils import get_chrome_driver

# 固定成都
CITY_CODE = "cd"
CITY_NAME = "成都"

# 室型URL映射
ROOM_MAP = {"1": "l1", "2": "l2", "3": "l3", "4": "l4", "5": "l5"}
_CN_NUM  = {"一": "1", "二": "2", "三": "3", "四": "4", "五": "5"}


def normalize_room(raw: str) -> str:
    """各种室型写法统一转为数字: 3/3室/三室/三室二厅 -> '3'"""
    if not raw:
        return ""
    for cn, num in _CN_NUM.items():
        raw = raw.replace(cn, num)
    m = re.search(r"([1-5])", raw)
    return m.group(1) if m else ""


def build_url(community: str, room: str, page: int = 1) -> str:
    """拼接贝壳网搜索URL，page>1时加pg段"""
    room_seg = ROOM_MAP.get(room, "")
    rs_seg   = f"rs{community}" if community else ""
    path_seg = f"{room_seg}{rs_seg}" if (room_seg or rs_seg) else ""
    pg_seg   = f"pg{page}/" if page > 1 else ""
    path     = f"/ershoufang/{path_seg}/{pg_seg}" if path_seg else f"/ershoufang/{pg_seg}"
    return f"https://{CITY_CODE}.ke.com{path}"


def parse_args():
    """解析命令行参数，返回 (community, room)"""
    args = sys.argv[1:]
    community, room = "", ""
    for arg in args:
        if arg.startswith("cs="):
            print("⚠ 当前仅支持成都，cs参数已忽略")
        elif not community:
            community = arg.strip()
        elif not room:
            room = normalize_room(arg.strip())
    return community, room


def download_image(url: str) -> bytes | None:
    """下载图片，带Referer防盗链头，失败返回None"""
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": config.USER_AGENT,
            "Referer": f"https://{CITY_CODE}.ke.com/"
        })
        return urllib.request.urlopen(req, timeout=10).read()
    except Exception:
        return None


def _is_match(community_input: str, community_result: str) -> bool:
    """
    判断搜索词与结果小区名是否匹配，支持三种情况：
    1. 完全包含：结果包含搜索词（鑫苑鑫都汇 in 鑫苑鑫都汇南区）
    2. 反向包含：搜索词包含结果（滨江和城一期 in 滨江和城）
    3. 关键词匹配：搜索词按2字切片，任意片段出现在结果中
       （滨江和城 -> [滨江, 江和, 和城] -> 和城 in 滨江和城二期）
    """
    if community_input in community_result:
        return True
    if community_result in community_input:
        return True
    # 2字滑动窗口切片匹配
    for i in range(len(community_input) - 1):
        chunk = community_input[i:i+2]
        if chunk in community_result:
            return True
    return False
    """下载图片，带Referer防盗链头，失败返回None"""
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": config.USER_AGENT,
            "Referer": f"https://{CITY_CODE}.ke.com/"
        })
        return urllib.request.urlopen(req, timeout=10).read()
    except Exception as e:
        print(f"  ⚠ 图片下载失败: {e}")
        return None


class BeikeScraper:
    """贝壳网成都二手房爬虫"""

    def __init__(self, community="", room=""):
        self.driver    = None
        self.data      = []
        self.community = community
        self.room      = room
        self.url       = build_url(community, room)

    def setup_driver(self):
        """初始化浏览器"""
        self.driver = get_chrome_driver(
            headless=config.HEADLESS,
            window_size=config.WINDOW_SIZE,
            user_agent=config.USER_AGENT
        )
        self.driver.implicitly_wait(config.IMPLICIT_WAIT)

    def open_page(self):
        """打开目标页面"""
        print(f"城市: {CITY_NAME}  小区: {self.community or '全部'}  室型: {self.room or '不限'}室")
        print(f"正在打开: {self.url}")
        self.driver.get(self.url)
        time.sleep(3)
        if "Forbidden" in self.driver.title:
            raise RuntimeError("页面加载异常（Forbidden），请稍后重试")
        print("✓ 页面加载完成")

    def sort_by_price(self):
        """点击总价排序"""
        try:
            btn = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), '总价')]"))
            )
            btn.click()
            time.sleep(2)
            print("✓ 已按总价排序")
        except Exception as e:
            print(f"⚠ 总价排序失败: {e}")

    def scrape_data(self, max_pages=1):
        """爬取房源列表，含懒加载触发、前5条验证、人机验证检测"""
        print(f"开始爬取，最多 {max_pages} 页...")

        for page in range(1, max_pages + 1):
            print(f"\n--- 第 {page} 页 ---")
            if page > 1:
                self.driver.get(build_url(self.community, self.room, page))
                time.sleep(random.uniform(3, 6))

            try:
                if "验证" in self.driver.title or "verify" in self.driver.current_url.lower():
                    print("⚠ 检测到人机验证，停止爬取")
                    break

                empty = self.driver.find_elements(By.CSS_SELECTOR, ".nullTip, .empty-content, .no-content")
                if empty:
                    print(f"⚠ 未找到房源：请确认小区名称「{self.community}」")
                    break

                items = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".sellListContent li.clear"))
                )

                # 分段滚动触发所有图片懒加载
                scroll_height = self.driver.execute_script("return document.body.scrollHeight")
                for step in range(0, scroll_height, 400):
                    self.driver.execute_script(f"window.scrollTo(0, {step});")
                    time.sleep(0.3)
                time.sleep(1)
                self.driver.execute_script("window.scrollTo(0, 0);")
                time.sleep(0.5)

                # 双重验证：新上房源通知弹窗 + 前5条均不匹配 -> 判定小区名输入有误
                if self.community and page == 1:
                    # 条件1：页面出现"新上房源通知"提示
                    has_notify = bool(self.driver.find_elements(
                        By.XPATH, "//div[contains(text(),'新上房源通知')]"
                    ))

                    # 条件2：前5条小区名均与输入无关
                    sample = []
                    for item in items[:5]:
                        try:
                            sample.append(item.find_element(By.CSS_SELECTOR, ".title a").text.split(" ")[0])
                        except:
                            pass
                    no_match = sample and not any(_is_match(self.community, n) for n in sample)

                    if has_notify and no_match:
                        room_tip = f"{self.room}室" if self.room else ""
                        print(f"⚠ 未找到「{self.community}」{room_tip}房源，请检查小区名称是否正确")
                        print(f"  推荐结果示例: {', '.join(sample[:3])}")
                        break

                for item in items:
                    try:
                        title_el   = item.find_element(By.CSS_SELECTOR, ".title a")
                        community  = title_el.text.strip().split(" ")[0]
                        link       = title_el.get_attribute("href") or ""
                        total_price = item.find_element(By.CSS_SELECTOR, ".totalPrice span").text.strip()
                        unit_price  = item.find_element(By.CSS_SELECTOR, ".unitPrice span").text.strip()

                        # 获取缩略图：优先取 data-original（懒加载真实URL），回退 src
                        img_url = ""
                        for img_el in item.find_elements(By.CSS_SELECTOR, "img.lj-lazy"):
                            src = img_el.get_attribute("data-original") or img_el.get_attribute("src") or ""
                            if src.startswith("http") and "blank.gif" not in src:
                                img_url = src
                                break

                        self.data.append({
                            "小区名称": community,
                            "室型":    f"{self.room}室" if self.room else "-",
                            "总价":    total_price + "万",
                            "单价":    unit_price,
                            "简图URL": img_url,
                            "链接":    link,
                            "爬取时间": time.strftime("%Y-%m-%d %H:%M:%S"),
                        })
                    except Exception:
                        continue

                print(f"✓ 本页 {len(items)} 条")

            except Exception as e:
                print(f"⚠ 第 {page} 页失败: {e}")
                break

        # 过滤不匹配的推荐结果（仅保留含搜索词的数据）
        if self.data and self.community:
            matched = [d for d in self.data if _is_match(self.community, d["小区名称"])]
            if matched:
                diff = len(self.data) - len(matched)
                if diff:
                    print(f"  过滤 {diff} 条不相关房源")
                self.data = matched
        if self.data:
            print(f"\n✓ 共爬取 {len(self.data)} 条")
        else:
            room_tip = f"{self.room}室" if self.room else ""
            print(f"\n✗ 无数据：请检查小区名称「{self.community}」或 成都 暂无{room_tip}房源")

    def save_data(self):
        """
        保存Excel，末行均价。
        图片策略：下载到 data/{小区名}/imgs/ 目录，Excel用相对路径引用。
        每次执行前清空该目录下的旧数据（表格+图片）。
        """
        if not self.data:
            print("⚠ 无数据可保存")
            return

        import shutil
        from openpyxl.utils import get_column_letter

        tag      = f"{CITY_NAME}_{self.community or '全部'}_{self.room or '不限'}室"
        out_dir  = os.path.join(config.OUTPUT_DIR, tag)   # data/成都_鑫苑鑫都汇_3室/
        img_dir  = os.path.join(out_dir, "imgs")          # data/成都_鑫苑鑫都汇_3室/imgs/

        # 每次执行先清空目录，保证数据最新
        if os.path.exists(out_dir):
            shutil.rmtree(out_dir)
            print(f"✓ 已清空旧数据: {out_dir}")
        os.makedirs(img_dir, exist_ok=True)

        df = pd.DataFrame(self.data)

        # 计算均价并追加汇总行
        df["_价值"] = df["单价"].str.extract(r"([\d,]+)").replace(",", "", regex=True).astype(float)
        avg = df["_价值"].mean()
        summary = {col: "" for col in df.columns}
        summary["小区名称"] = "均价"
        summary["单价"]    = f"{avg:,.0f}元/平"
        df = pd.concat([df, pd.DataFrame([summary])], ignore_index=True)
        df = df.drop(columns=["_价值", "简图URL"])

        # 并发下载图片到本地
        from concurrent.futures import ThreadPoolExecutor, as_completed
        img_urls = [d.get("简图URL", "") for d in self.data] + [""]
        valid_urls = {i: u for i, u in enumerate(img_urls) if u}
        print(f"正在并发下载 {len(valid_urls)} 张图片...")

        img_files = {}  # idx -> 本地文件名
        def _download(idx, url):
            data = download_image(url)
            if data:
                fname = f"{idx+1}.jpg"
                fpath = os.path.join(img_dir, fname)
                with open(fpath, "wb") as f:
                    f.write(data)
                return idx, fname
            return idx, None

        with ThreadPoolExecutor(max_workers=8) as executor:
            futures = {executor.submit(_download, idx, url): idx for idx, url in valid_urls.items()}
            for future in as_completed(futures):
                idx, fname = future.result()
                if fname:
                    img_files[idx] = fname

        # 写入Excel，简图列用相对路径文本（可在Excel中手动查看）
        # 同时在简图列旁边插入实际图片
        filepath = os.path.join(out_dir, f"beike_{tag}.xlsx")
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)
            ws = writer.sheets["Sheet1"]

            # 链接列超链接
            link_col = df.columns.get_loc("链接") + 1
            for r, val in enumerate(df["链接"], start=2):
                if val and str(val).startswith("http"):
                    ws.cell(row=r, column=link_col).hyperlink = val
                    ws.cell(row=r, column=link_col).style = "Hyperlink"

            # 简图列：插入本地图片
            from openpyxl.drawing.image import Image as XLImage
            img_col = len(df.columns) + 1
            ws.cell(row=1, column=img_col).value = "简图"
            ws.column_dimensions[get_column_letter(img_col)].width = 22

            for idx, fname in img_files.items():
                r = idx + 2
                ws.row_dimensions[r].height = 80
                fpath = os.path.join(img_dir, fname)
                try:
                    img_obj = XLImage(fpath)
                    img_obj.width, img_obj.height = 120, 90
                    ws.add_image(img_obj, f"{get_column_letter(img_col)}{r}")
                except Exception:
                    pass

        print(f"✓ 已保存: {filepath}")
        print(f"  图片目录: {img_dir}  (共 {len(img_files)} 张)")
        print(f"  均价: {avg:,.0f}元/平")

    def close(self):
        if self.driver:
            self.driver.quit()
            print("✓ 浏览器已关闭")

    def run(self, max_pages=1):
        try:
            self.setup_driver()
            self.open_page()
            self.sort_by_price()
            self.scrape_data(max_pages)
            self.save_data()
        except Exception as e:
            print(f"✗ 出错: {e}")
        finally:
            self.close()


if __name__ == "__main__":
    community, room = parse_args()
    if not community:
        print("用法: python scrapers/beike_scraper.py <小区名> [几室]")
        print("示例: python scrapers/beike_scraper.py 鑫苑鑫都汇 3")
        sys.exit(0)
    BeikeScraper(community=community, room=room).run(max_pages=1)
